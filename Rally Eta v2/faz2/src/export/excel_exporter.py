"""
Excel export for rally predictions.

Supports:
- Single prediction export
- Batch prediction export (multiple drivers)
- Detailed explanation sheets
- Summary sheets for rally officials
"""
import logging
from typing import List, Optional
from datetime import datetime
from pathlib import Path
from collections import defaultdict
from dataclasses import dataclass
from zipfile import ZipFile, ZIP_DEFLATED
from xml.sax.saxutils import escape

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

    @dataclass
    class Font:
        bold: bool = False
        size: Optional[int] = None
        color: Optional[str] = None
        name: Optional[str] = None

    @dataclass
    class PatternFill:
        start_color: Optional[str] = None
        end_color: Optional[str] = None
        fill_type: Optional[str] = None

    @dataclass
    class Alignment:
        horizontal: Optional[str] = None

    @dataclass
    class Side:
        style: Optional[str] = None

    @dataclass
    class Border:
        left: Optional[Side] = None
        right: Optional[Side] = None
        top: Optional[Side] = None
        bottom: Optional[Side] = None

    def get_column_letter(column_index: int) -> str:
        """Convert 1-based column index to Excel letter."""
        result = []
        while column_index:
            column_index, remainder = divmod(column_index - 1, 26)
            result.append(chr(65 + remainder))
        return ''.join(reversed(result))

    @dataclass
    class _FallbackCell:
        value: Optional[object] = None
        font: Optional[Font] = None
        fill: Optional[PatternFill] = None
        alignment: Optional[Alignment] = None
        border: Optional[Border] = None

    @dataclass
    class _FallbackColumnDimension:
        width: Optional[float] = None

    class _FallbackColumnDimensions(defaultdict):
        def __init__(self):
            super().__init__(_FallbackColumnDimension)

    class _FallbackWorksheet:
        def __init__(self, title: str = "Sheet"):
            self.title = title
            self._cells = {}
            self._merged_ranges = []
            self.column_dimensions = _FallbackColumnDimensions()

        def merge_cells(self, cell_range: str):
            self._merged_ranges.append(cell_range)

        def cell(self, row: int, column: int, value=None):
            ref = f"{get_column_letter(column)}{row}"
            cell = self._cells.setdefault(ref, _FallbackCell())
            if value is not None:
                cell.value = value
            return cell

        def __getitem__(self, ref: str):
            return self._cells.setdefault(ref, _FallbackCell())

        def __setitem__(self, ref: str, value):
            cell = self.__getitem__(ref)
            cell.value = value

    class Workbook:
        """Small XLSX fallback writer for test/sandbox environments."""

        def __init__(self):
            self._worksheets = [_FallbackWorksheet()]
            self.active = self._worksheets[0]

        def create_sheet(self, title: str):
            ws = _FallbackWorksheet(title)
            self._worksheets.append(ws)
            return ws

        def save(self, output_path: str):
            output_file = Path(output_path)
            output_file.parent.mkdir(parents=True, exist_ok=True)
            sheet_entries = []
            sheet_rel_entries = []

            for index, worksheet in enumerate(self._worksheets, start=1):
                sheet_entries.append(
                    f'<sheet name="{escape(worksheet.title)}" sheetId="{index}" r:id="rId{index}"/>'
                )
                sheet_rel_entries.append(
                    f'<Relationship Id="rId{index}" '
                    f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
                    f'Target="worksheets/sheet{index}.xml"/>'
                )

            workbook_xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
                'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
                f'<sheets>{"".join(sheet_entries)}</sheets>'
                '</workbook>'
            )

            workbook_rels_xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                f'{"".join(sheet_rel_entries)}'
                '</Relationships>'
            )

            content_types = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
                '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
                '<Default Extension="xml" ContentType="application/xml"/>'
                '<Override PartName="/xl/workbook.xml" '
                'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
                '<Override PartName="/docProps/core.xml" '
                'ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>'
                '<Override PartName="/docProps/app.xml" '
                'ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>'
                + ''.join(
                    f'<Override PartName="/xl/worksheets/sheet{idx}.xml" '
                    'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
                    for idx in range(1, len(self._worksheets) + 1)
                )
                + '</Types>'
            )

            rels_xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rId1" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
                'Target="xl/workbook.xml"/>'
                '<Relationship Id="rId2" '
                'Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" '
                'Target="docProps/core.xml"/>'
                '<Relationship Id="rId3" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" '
                'Target="docProps/app.xml"/>'
                '</Relationships>'
            )

            core_xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
                'xmlns:dc="http://purl.org/dc/elements/1.1/" '
                'xmlns:dcterms="http://purl.org/dc/terms/" '
                'xmlns:dcmitype="http://purl.org/dc/dcmitype/" '
                'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
                '<dc:creator>Codex</dc:creator>'
                '<cp:lastModifiedBy>Codex</cp:lastModifiedBy>'
                '</cp:coreProperties>'
            )

            app_xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" '
                'xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">'
                '<Application>Codex</Application>'
                '</Properties>'
            )

            with ZipFile(output_file, 'w', ZIP_DEFLATED) as zf:
                zf.writestr("[Content_Types].xml", content_types)
                zf.writestr("_rels/.rels", rels_xml)
                zf.writestr("docProps/core.xml", core_xml)
                zf.writestr("docProps/app.xml", app_xml)
                zf.writestr("xl/workbook.xml", workbook_xml)
                zf.writestr("xl/_rels/workbook.xml.rels", workbook_rels_xml)

                for index, worksheet in enumerate(self._worksheets, start=1):
                    zf.writestr(f"xl/worksheets/sheet{index}.xml", _build_sheet_xml(worksheet))

    def _build_sheet_xml(worksheet: _FallbackWorksheet) -> str:
        rows = defaultdict(list)
        for ref, cell in worksheet._cells.items():
            coord = _parse_cell_ref(ref)
            rows[coord[1]].append((coord[0], ref, cell.value))

        row_xml = []
        for row_number in sorted(rows):
            cells_xml = []
            for _, ref, value in sorted(rows[row_number], key=lambda item: item[0]):
                cells_xml.append(_build_cell_xml(ref, value))
            row_xml.append(f'<row r="{row_number}">{"".join(cells_xml)}</row>')

        merge_xml = ""
        if worksheet._merged_ranges:
            merge_entries = ''.join(
                f'<mergeCell ref="{escape(cell_range)}"/>' for cell_range in worksheet._merged_ranges
            )
            merge_xml = f'<mergeCells count="{len(worksheet._merged_ranges)}">{merge_entries}</mergeCells>'

        return (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            f'<sheetData>{"".join(row_xml)}</sheetData>'
            f'{merge_xml}'
            '</worksheet>'
        )

    def _build_cell_xml(ref: str, value) -> str:
        if value is None:
            return f'<c r="{ref}"/>'

        if isinstance(value, bool):
            return f'<c r="{ref}" t="b"><v>{int(value)}</v></c>'

        if isinstance(value, (int, float)):
            return f'<c r="{ref}"><v>{value}</v></c>'

        text = escape(str(value))
        return f'<c r="{ref}" t="inlineStr"><is><t>{text}</t></is></c>'

    def _parse_cell_ref(ref: str):
        letters = ''.join(ch for ch in ref if ch.isalpha()).upper()
        digits = ''.join(ch for ch in ref if ch.isdigit())

        col_num = 0
        for char in letters:
            col_num = (col_num * 26) + (ord(char) - 64)

        return col_num, int(digits)

logger = logging.getLogger(__name__)

if not OPENPYXL_AVAILABLE:
    logger.warning("openpyxl not available, using fallback XLSX writer")


# Style definitions
STYLES = {
    'header': Font(bold=True, size=14, color="FFFFFF"),
    'header_fill': PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid"),
    'section_header': Font(bold=True, size=12),
    'section_fill': PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid"),
    'prediction': Font(bold=True, size=14, color="C00000"),
    'confidence_high': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    'confidence_medium': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    'confidence_low': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    'border': Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
}


class ExcelExporter:
    """
    Export predictions to Excel.

    Supports both single and batch exports with
    detailed explanations.
    """

    def __init__(self):
        """Initialize exporter."""
        pass

    def export_prediction(
        self,
        prediction,  # PredictionResult
        output_path: str,
        include_details: bool = True
    ):
        """
        Export single prediction to Excel.

        Args:
            prediction: PredictionResult object
            output_path: Output Excel file path
            include_details: Whether to include detailed explanation sheet
        """
        wb = Workbook()

        # Main prediction sheet
        self._create_single_prediction_sheet(wb, prediction)

        # Detailed explanation sheet
        if include_details:
            self._create_explanation_sheet(wb, prediction)

        # Save
        wb.save(output_path)
        logger.info(f"Excel exported to: {output_path}")

    def export_batch(
        self,
        predictions: List,  # List[PredictionResult]
        rally_name: str,
        stage_name: str,
        output_path: str,
        include_details: bool = True
    ):
        """
        Export batch predictions to Excel.

        Args:
            predictions: List of PredictionResult objects
            rally_name: Rally name for title
            stage_name: Stage name for title
            output_path: Output Excel file path
            include_details: Whether to include detailed sheets
        """
        wb = Workbook()

        # Summary sheet (main)
        self._create_summary_sheet(wb, predictions, rally_name, stage_name)

        # Individual prediction sheets if requested
        if include_details:
            for i, pred in enumerate(predictions[:10], 1):  # Limit to 10 detail sheets
                self._create_driver_sheet(wb, pred, i)

        # Save
        wb.save(output_path)
        logger.info(f"Batch Excel exported to: {output_path}")

    def _create_single_prediction_sheet(self, wb: Workbook, prediction):
        """Create main prediction sheet for single export."""
        ws = wb.active
        ws.title = "Tahmin"

        row = 1

        # Title
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "Rally ETA v2.0 - Notional Time Tahmini"
        ws[f'A{row}'].font = STYLES['header']
        ws[f'A{row}'].fill = STYLES['header_fill']
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2

        # Driver & Stage Info
        ws[f'A{row}'] = "Pilot:"
        ws[f'B{row}'] = prediction.driver_name
        ws[f'B{row}'].font = Font(bold=True)
        row += 1

        ws[f'A{row}'] = "Etap:"
        ws[f'B{row}'] = prediction.stage_name
        row += 1

        ws[f'A{row}'] = "Sinif:"
        ws[f'B{row}'] = prediction.normalized_class
        row += 1

        ws[f'A{row}'] = "Zemin:"
        ws[f'B{row}'] = prediction.surface.capitalize()
        row += 2

        # Prediction Section
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "TAHMIN SONUCU"
        ws[f'A{row}'].font = STYLES['section_header']
        ws[f'A{row}'].fill = STYLES['section_fill']
        row += 1

        ws[f'A{row}'] = "Tahmini Zaman:"
        ws[f'B{row}'] = prediction.predicted_time_str
        ws[f'B{row}'].font = STYLES['prediction']
        row += 1

        ws[f'A{row}'] = "Final Ratio:"
        ws[f'B{row}'] = f"{prediction.predicted_ratio:.4f}"
        row += 1

        ws[f'A{row}'] = "Sinif Lideri:"
        ws[f'B{row}'] = f"{prediction.class_best_driver} ({prediction.class_best_str})"
        row += 1

        diff_seconds = (prediction.predicted_ratio - 1) * prediction.class_best_time
        ws[f'A{row}'] = "Fark:"
        ws[f'B{row}'] = f"+{diff_seconds:.1f} saniye"
        row += 2

        # Components Section
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "HESAPLAMA DETAYLARI"
        ws[f'A{row}'].font = STYLES['section_header']
        ws[f'A{row}'].fill = STYLES['section_fill']
        row += 1

        ws[f'A{row}'] = "Baseline Ratio:"
        ws[f'B{row}'] = f"{prediction.baseline_ratio:.4f}"
        row += 1

        ws[f'A{row}'] = "Momentum Factor:"
        ws[f'B{row}'] = f"{prediction.momentum_factor:.4f}"
        row += 1

        ws[f'A{row}'] = "Surface Adjustment:"
        ws[f'B{row}'] = f"{prediction.surface_adjustment:.4f}"
        row += 1

        ws[f'A{row}'] = "Geometrik Duzeltme:"
        ws[f'B{row}'] = f"{prediction.geometric_correction:.4f}"
        ws[f'C{row}'] = f"({prediction.geometric_mode})"
        row += 2

        # Confidence Section
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "GUVENILIRLIK"
        ws[f'A{row}'].font = STYLES['section_header']
        ws[f'A{row}'].fill = STYLES['section_fill']
        row += 1

        ws[f'A{row}'] = "Guven Seviyesi:"
        ws[f'B{row}'] = f"{prediction.confidence.level} ({prediction.confidence.score}/100)"

        # Color code confidence
        if prediction.confidence.level == 'HIGH':
            ws[f'B{row}'].fill = STYLES['confidence_high']
        elif prediction.confidence.level == 'MEDIUM':
            ws[f'B{row}'].fill = STYLES['confidence_medium']
        else:
            ws[f'B{row}'].fill = STYLES['confidence_low']
        row += 2

        # Confidence reasons
        for reason in prediction.confidence.reasons:
            if reason:
                ws[f'A{row}'] = f"  - {reason}"
                row += 1

        row += 1

        # Timestamp
        ws[f'A{row}'] = "Olusturma Tarihi:"
        ws[f'B{row}'] = prediction.generated_at[:19].replace('T', ' ')

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def _create_explanation_sheet(self, wb: Workbook, prediction):
        """Create detailed explanation sheet."""
        ws = wb.create_sheet("Detayli Aciklama")

        # Split detailed text into lines
        lines = prediction.detailed_text.split('\n')

        for i, line in enumerate(lines, 1):
            ws[f'A{i}'] = line
            ws[f'A{i}'].font = Font(name='Consolas', size=10)

        ws.column_dimensions['A'].width = 80

    def _create_summary_sheet(
        self,
        wb: Workbook,
        predictions: List,
        rally_name: str,
        stage_name: str
    ):
        """Create summary sheet for batch export."""
        ws = wb.active
        ws.title = "Ozet"

        row = 1

        # Title
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = f"Rally ETA v2.0 - {rally_name} / {stage_name}"
        ws[f'A{row}'].font = STYLES['header']
        ws[f'A{row}'].fill = STYLES['header_fill']
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1

        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = f"Olusturma: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2

        # Header row
        headers = [
            'Pilot', 'Sinif', 'Tahmini Zaman', 'Ratio',
            'Baseline', 'Geo. Duz.', 'Guven', 'Fark (sn)'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = STYLES['section_fill']
            cell.border = STYLES['border']
            cell.alignment = Alignment(horizontal='center')

        row += 1

        # Data rows
        for pred in predictions:
            diff_seconds = (pred.predicted_ratio - 1) * pred.class_best_time

            data = [
                pred.driver_name,
                pred.normalized_class,
                pred.predicted_time_str,
                f"{pred.predicted_ratio:.3f}",
                f"{pred.baseline_ratio:.3f}",
                f"{pred.geometric_correction:.3f}",
                pred.confidence.level,
                f"+{diff_seconds:.1f}"
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = STYLES['border']

                # Center align numeric columns
                if col >= 3:
                    cell.alignment = Alignment(horizontal='center')

                # Color code confidence
                if col == 7:
                    if value == 'HIGH':
                        cell.fill = STYLES['confidence_high']
                    elif value == 'MEDIUM':
                        cell.fill = STYLES['confidence_medium']
                    else:
                        cell.fill = STYLES['confidence_low']

            row += 1

        # Column widths
        widths = [25, 10, 15, 10, 10, 10, 10, 12]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Add summary statistics
        row += 2
        ws[f'A{row}'] = "ISTATISTIKLER"
        ws[f'A{row}'].font = STYLES['section_header']
        row += 1

        ws[f'A{row}'] = f"Toplam Pilot: {len(predictions)}"
        row += 1

        if predictions:
            avg_ratio = sum(p.predicted_ratio for p in predictions) / len(predictions)
            ws[f'A{row}'] = f"Ortalama Ratio: {avg_ratio:.3f}"
            row += 1

            high_conf = sum(1 for p in predictions if p.confidence.level == 'HIGH')
            med_conf = sum(1 for p in predictions if p.confidence.level == 'MEDIUM')
            low_conf = len(predictions) - high_conf - med_conf

            ws[f'A{row}'] = f"Guven Dagilimi: HIGH={high_conf}, MEDIUM={med_conf}, LOW={low_conf}"

    def _create_driver_sheet(self, wb: Workbook, prediction, index: int):
        """Create individual driver sheet."""
        # Truncate name for sheet title (max 31 chars)
        sheet_name = f"{index}_{prediction.driver_name[:25]}"
        ws = wb.create_sheet(sheet_name)

        # Use same format as single prediction
        row = 1

        ws[f'A{row}'] = f"PILOT: {prediction.driver_name}"
        ws[f'A{row}'].font = STYLES['section_header']
        row += 2

        ws[f'A{row}'] = f"Tahmini Zaman: {prediction.predicted_time_str}"
        ws[f'A{row}'].font = STYLES['prediction']
        row += 1

        ws[f'A{row}'] = f"Final Ratio: {prediction.predicted_ratio:.4f}"
        row += 1

        ws[f'A{row}'] = f"Guven: {prediction.confidence.level} ({prediction.confidence.score}/100)"
        row += 2

        # Detailed explanation
        ws[f'A{row}'] = "DETAYLI ACIKLAMA"
        ws[f'A{row}'].font = STYLES['section_header']
        row += 1

        lines = prediction.detailed_text.split('\n')
        for line in lines:
            ws[f'A{row}'] = line
            ws[f'A{row}'].font = Font(name='Consolas', size=9)
            row += 1

        ws.column_dimensions['A'].width = 70


# Legacy support for old format
def export_prediction(prediction: dict, driver_name: str,
                     stage_name: str, output_path: str):
    """
    Legacy export function for backward compatibility.

    Args:
        prediction: Prediction dict from BaselinePredictor
        driver_name: Driver name
        stage_name: Stage name
        output_path: Output Excel file path
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Tahmin"

    # Header
    ws['A1'] = "Rally ETA v2.0 - Notional Time Tahmini"
    ws['A1'].font = Font(bold=True, size=14)

    # Driver & Stage
    ws['A3'] = "Pilot:"
    ws['B3'] = driver_name
    ws['A4'] = "Etap:"
    ws['B4'] = stage_name

    # Prediction
    ws['A6'] = "TAHMIN"
    ws['A6'].font = Font(bold=True, size=12)

    ws['A7'] = "Tahmin Edilen Zaman:"
    ws['B7'] = prediction['predicted_time_str']
    ws['B7'].font = Font(bold=True, size=12, color="FF0000")

    ws['A8'] = "Ratio:"
    ws['B8'] = f"{prediction['baseline_ratio']:.3f}"

    ws['A9'] = "Sinif Lideri:"
    ws['B9'] = prediction.get('class_best_driver', 'N/A')

    ws['A10'] = "Sinif Lideri Zamani:"
    ws['B10'] = prediction.get('class_best_str', 'N/A')

    ws['A11'] = "Fark:"
    diff = prediction.get('predicted_time', 0) - prediction.get('class_best_time', 0)
    ws['B11'] = f"+{diff:.1f} saniye"

    # Components
    ws['A13'] = "HESAPLAMA DETAYLARI"
    ws['A13'].font = Font(bold=True)

    components = prediction.get('components', {})
    ws['A14'] = "Driver Baseline:"
    ws['B14'] = f"{components.get('driver_baseline', prediction['baseline_ratio']):.3f}"

    ws['A15'] = "Momentum Factor:"
    ws['B15'] = f"{components.get('momentum_factor', 1.0):.3f}"

    ws['A16'] = "Surface Adjustment:"
    ws['B16'] = f"{components.get('surface_adjustment', 1.0):.3f}"

    ws['A18'] = "Confidence:"
    ws['B18'] = prediction.get('confidence', 'N/A')

    # Formatting
    for row_num in [1, 6, 13]:
        ws[f'A{row_num}'].fill = PatternFill(start_color="CCCCCC", fill_type="solid")

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20

    # Save
    wb.save(output_path)
    logger.info(f"Excel exported to: {output_path}")


def main():
    """Test Excel exporter."""
    from dataclasses import dataclass
    from typing import List, Dict

    # Create mock prediction result
    @dataclass
    class MockConfidence:
        level: str = "HIGH"
        score: int = 85
        emoji: str = "🟢"
        reasons: List[str] = None
        breakdown: Dict[str, int] = None

        def __post_init__(self):
            if self.reasons is None:
                self.reasons = ["Yeterli gecmis veri", "Surface deneyimi var"]
            if self.breakdown is None:
                self.breakdown = {'historical': 40, 'surface': 25}

    @dataclass
    class MockPrediction:
        driver_id: str = "test_driver"
        driver_name: str = "Test Pilot"
        stage_id: str = "SS3"
        stage_name: str = "SS3 - Catalca"
        normalized_class: str = "Rally2"
        surface: str = "gravel"
        predicted_time_seconds: float = 645.5
        predicted_time_str: str = "10:45.500"
        predicted_ratio: float = 1.052
        class_best_time: float = 613.5
        class_best_str: str = "10:13.500"
        class_best_driver: str = "Sinif Lideri"
        baseline_ratio: float = 1.045
        momentum_factor: float = 1.01
        surface_adjustment: float = 0.98
        geometric_correction: float = 1.015
        geometric_mode: str = "geometric"
        confidence: MockConfidence = None
        summary_text: str = "Test summary"
        detailed_text: str = "Test detailed explanation\n" * 10
        generated_at: str = "2025-12-29T10:00:00"

        def __post_init__(self):
            if self.confidence is None:
                self.confidence = MockConfidence()

    print("Testing Excel Exporter...")

    exporter = ExcelExporter()

    # Test single export
    pred = MockPrediction()
    exporter.export_prediction(pred, "test_single.xlsx")
    print("Single export: test_single.xlsx")

    # Test batch export
    predictions = [
        MockPrediction(driver_name=f"Pilot {i}", predicted_ratio=1.05 + i*0.01)
        for i in range(5)
    ]

    exporter.export_batch(
        predictions,
        rally_name="Test Rally",
        stage_name="SS3",
        output_path="test_batch.xlsx"
    )
    print("Batch export: test_batch.xlsx")

    print("\nExport tests completed!")


if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO)
    main()
